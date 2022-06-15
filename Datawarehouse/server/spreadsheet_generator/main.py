import os
from io import BytesIO

from typing import Optional
from fastapi import FastAPI, File, Request, status
from pydantic import BaseModel
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse, Response

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class DataCodeModel(BaseModel):
    # barcode values go here
    rows: list
    # identify caller (android device, browser, hostname etc.)
    caller: str
    # filename for spreadsheet
    filename: str
    # if a header row is present, pass as True
    has_header: bool


def generate_spreadsheet (rows, has_header, filename):
    wb = Workbook()
    ws = wb.active

    # calculate correct width for each column
    # based on how long the longest string for each specific column is
    cell_width_calculation = {}
    for i in range(len(rows[0])):
        cell_width_calculation[i] = 0
    for index in cell_width_calculation:
        for row in rows:
            if row[index] is not None:
                if len(str(row[index])) > cell_width_calculation[index]:
                    cell_width_calculation[index] = len(str(row[index]))
    for key in cell_width_calculation:
        asccii_number = key + 65 # get A, B, C .... using numbers
        column = chr(asccii_number)
        print(column)
        ws.column_dimensions[column].width = cell_width_calculation[key]+5

    # append rows to worksheet
    for row in rows:
        ws.append(row)

    # freeze scroll for keeping headers visible
    if has_header:
        ws.freeze_panes = ws['A2']

    byte_array = BytesIO()
    wb.save(byte_array)
    byte_array.seek(0)
    # about seek(0) -> https://docs.python.org/3/library/io.html#io.IOBase.seek
    return byte_array


app = FastAPI()

@app.post("/spreadsheet")
async def shelf_multiple(item: DataCodeModel):
    print('creating spreadsheet for ' + item.caller)
    print(item.filename)
    byte_array = generate_spreadsheet(item.rows, item.has_header, item.filename)
    headers = {
        'Content-Disposition': 'attachment; filename="' + item.filename + '.xlsx"'
    }
    return StreamingResponse(byte_array, status_code=201, headers=headers)
