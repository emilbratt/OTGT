#!/usr/bin/env python3
__author__ = "Emil Bratt Børsting"
__email__  = "emilbratt@gmail.com"
__status__ = "developement"
import os
import sys
import csv
import json
import subprocess
from dataget import Getconnect
from datapost import Postconnect
from datetime import datetime
from credentials import loadCredentials
from writelog import Log

# get timestamp and date
timestamp = datetime.now()
ymd = timestamp.strftime("%Y-%m-%d")

# set working directory
dirs = {}
dirs['App'] = os.path.dirname(os.path.realpath(__file__))
dirs['Data'] = os.path.join(dirs['App'], 'Data')
dirs['Utsolgt'] = os.path.join(dirs['App'], 'Data','Utsolgt')
dirs['Import'] = os.path.join(dirs['App'], 'Data', 'Import')
dirs['Omsetning'] = os.path.join(dirs['App'], 'Data', 'Omsetning')
dirs['Salg'] = os.path.join(dirs['App'], 'Data', 'Salg')

# if no dir, create new
os.makedirs(dirs['Data'], exist_ok=True)
os.makedirs(dirs['Utsolgt'], exist_ok=True)
os.makedirs(dirs['Import'], exist_ok=True)
os.makedirs(dirs['Omsetning'], exist_ok=True)
os.makedirs(dirs['Utsolgt'], exist_ok=True)
os.makedirs(dirs['Salg'], exist_ok=True)
for dir in dirs:
    if dir != 'Data' and dir != 'App':
        os.makedirs(dirs[dir]+'/Daglig', exist_ok=True)
        if dir != 'Salg':
            os.makedirs(dirs[dir]+'/Ukentlig', exist_ok=True)
            os.makedirs(dirs[dir]+'/Maanedlig', exist_ok=True)


def initialize():

    P = Postconnect()
    G = Getconnect()

    P.makeTables() # if tables for data warehouse do not exist, create them

    # get datetimes and check if weekly and/or monthly execute = True
    data['Tider'] = G.fetchTime()


    # get highest id number for article id and brand id
    brand_idMax = P.brandsGetMax()
    article_idMax = P.articlesGetMax()


    # fetch new brands and articles from retial
    brandsPost = []
    for row in G.getBrands(brand_idMax):
        brandsPost.append(tuple(row))
    if brandsPost != []:
        Log(f'Updating brands from brand_id: {brand_idMax}')
        P.brandsPost(brandsPost)
    else:
        Log(f'No new brands from brand_id: {brand_idMax}')

    articlesPost = []
    for row in G.getArticles(article_idMax):
        articlesPost.append(tuple(row))
    if articlesPost != []:
        Log(f'Updating articles from article_id: {article_idMax}')
        P.articlesPost(articlesPost)
    else:
        Log(f'No new articles from article_id: {article_idMax}')


    # update barcodes
    if data['Tider']['monthly'] == True:
        barcodes = G.getBarcodes()
        '''
        fetch new barcodes from retial
        barcodes can be changed, as an easy workaround,
        instead of updating incrementally we have to delete
        all records and insert all new barcodes over again
        '''
        barcodePost = []
        for row in barcodes:
            barcodePost.append(tuple(row))
        P.barcodesDel()
        P.barcodesPost(barcodePost)

    G.close()
    P.close()


def getRecords():
    c = Getconnect()

    # # daily
    data['Omsetning'] = {'Daglig':c.turnoverDaily()}
    data['Import'] = {'Daglig':c.importsDaily()}
    data['Utsolgt'] = {'Daglig':c.soldoutDaily()}
    data['Salg'] = {'Daglig':c.salesDaily()}
    post['sales_count'] = {'daily':c.salesCountDaily()}

    # weekly
    if data['Tider']['weekly'] == True:
        # onlyt run on mondays (get last week)
        data['Omsetning']['Ukentlig'] = c.turnoverWeekly()
        data['Utsolgt']['Ukentlig'] = c.soldoutWeekly()
        data['Import']['Ukentlig'] = c.importsWeekly()

    # monthly
    if data['Tider']['monthly'] == True:
         # only run on first every month (get last month)
        data['Omsetning']['Maanedlig'] = c.turnoverMonthly()
        data['Utsolgt']['Maanedlig'] = c.soldoutMonthly()
        data['Import']['Maanedlig'] = c.importsMonthly()

    c.close()


def updateCIP():
     # for rows that need adjustments
    prepared = {}
    # format dates for columns in data warehouse
    dateInsert = []
    dateInsert.append(int(data['Tider']['yesterday']['YYYYMMDD'][:4]))
    dateInsert.append(int(data['Tider']['yesterday']['YYYYMMDD'][4:6]))
    dateInsert.append(int(data['Tider']['yesterday']['YYYYMMDD'][6:8]))
    dateInsert.append(int(data['Tider']['yesterday']['weekNum']))
    dateInsert.append(data['Tider']['yesterday']['weekday'])
    dateInsert.append(int(data['Tider']['yesterday']['YYYYMMDD']))
    dateInsert.append(data['Tider']['yesterday']['human'])


    def prepareRecords(row,category):
        if category == 'Omsetning':
            turnover = {}
            # extract hourly values and add dates
            turnover['timer'] = [v for v in row[1:]]
            for value in dateInsert:
                turnover['timer'].append(value)
            # extract total value and dates
            turnover['dag'] = [row[0]]
            for value in dateInsert:
                turnover['dag'].append(value)
            return turnover

        elif category == 'Utsolgt' or category == 'Import':
            for value in dateInsert:
                row.append(value)
            return row
        elif category == 'Salg':
            return row
        elif category == 'sales_count':
            add = []
            add = [v for v in row]
            for value in dateInsert:
                post[category][when].append(value)
                add.append(value)
            return add
        else:
            return None


    for category in data:
        if category != 'Tider':
            for when in data[category]:
                if when == 'Daglig':
                    prepared[category] = [] # add category as key
                    # prepare rows with additional data
                    for row in data[category][when]:
                        prepared[category].append(
                            prepareRecords(list(row),category))

    for category in post:
        for when in post[category]:
            prepareRecords(post[category][when],category)

    # insert rows in data warehouse
    c = Postconnect()
    c.soldoutPost(prepared['Utsolgt'])
    c.importsPost(prepared['Import'])
    c.salesPost(prepared['Salg'])
    c.turnover_hourlyPost(prepared['Omsetning'][0]['timer'])
    c.turnover_dailyPost(prepared['Omsetning'][0]['dag'])
    c.sales_countPost(post['sales_count']['daily'])

    c.close()


def writeSpreadsheet():
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    credentials = loadCredentials('cloud')

    def exportXLSX(fileName: str):


        def cloudUpload(fileName: str):
            def commandAsmbl(cmd):
                Log(f'Cloud Upload: uploading {category} {when} to {credentials["server"]}')
                subprocess.Popen(cmd, shell=True, executable='/bin/bash')

            commmand = 'curl -u '
            commmand += credentials['user']
            commmand += ':'
            commmand += credentials['password']
            commmand += ' -T '
            commmand += fileName
            commmand += ' https://'
            commmand += credentials['server']
            commmand += '/remote.php/dav/files/'
            commmand += credentials['user']
            commmand += '/'
            commmand += category
            commmand += '/'
            commmand += when
            commmand += '/'
            commandAsmbl(commmand)


        # load spreadsheet
        wb = Workbook()
        ws = wb.active

        # apply default col length
        cellLength = {}
        for i in range(10):
            cellLength[i] = 5

        # set length of col based on length of longest cell value
        for row in data[category][when]:
            for i,cell in enumerate(row):
                if i not in cellLength:
                    cellLength[i] = 1
                try:
                    if cellLength[i] < len(str(cell)):
                        cellLength[i] = len(str(cell))
                except KeyError:
                    continue

        # apply col length from values in cellLength to cell A-J
        ws.column_dimensions['A'].width = cellLength[0]+5
        ws.column_dimensions['B'].width = cellLength[1]+5
        ws.column_dimensions['C'].width = cellLength[2]+5
        ws.column_dimensions['D'].width = cellLength[3]+5
        ws.column_dimensions['E'].width = cellLength[4]+5
        ws.column_dimensions['F'].width = cellLength[5]+5
        ws.column_dimensions['G'].width = cellLength[6]+5
        ws.column_dimensions['H'].width = cellLength[7]+5
        ws.column_dimensions['I'].width = cellLength[8]+5
        ws.column_dimensions['J'].width = cellLength[9]+5

        freeze = None  # find freeze point for titles while appending
        for i,row in enumerate(data[category][when]):
            if len(row) > 2 and freeze == None:
                freeze = 'A'+str(i+2)
            if category == 'Salg' and i > 2:
                ws.append(list(row)[:-1])
            else:
                ws.append(list(row))

        for col in ws.columns: # align all fields with value to center
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        if freeze != None:
            ws.freeze_panes = ws[freeze]

        # save spreadsheet
        wb.save(fileName)

        # send wb to cloud
        cloudUpload(fileName)


    # give column names and add date info before exporting to spreadsheet

    colName = [
    'Totalt','00-01','01-02','02-03','03-04','04-05','05-06','06-07','07-08',
    '08-09','09-10','10-11','11-12','12-13','13-14','14-15','15-16','16-17',
    '17-18','18-19','19-20','20-21','21-22','22-23','23-24'
    ]
    data['Omsetning']['Daglig'].insert(0,colName)
    data['Omsetning']['Daglig'].insert(0,[data['Tider']['yesterday']['weekday'].title() +' Uke-' + data['Tider']['yesterday']['weekNum']])
    data['Omsetning']['Daglig'].insert(0,[data['Tider']['yesterday']['human']])
    data['Omsetning']['Daglig'].insert(0,['Omsetning'])

    colName = [
    'Artikkel ID','Merke','Navn',
    'Importert','Antall Lager','Lagerplass','Lev.ID'
    ]
    data['Import']['Daglig'].insert(0,colName)
    data['Import']['Daglig'].insert(0,[data['Tider']['yesterday']['weekday'].title() +' Uke-' + data['Tider']['yesterday']['weekNum']])
    data['Import']['Daglig'].insert(0,[data['Tider']['yesterday']['human']])
    data['Import']['Daglig'].insert(0,['Vareimport'])

    colName = [
    'Artikkel ID','Merke','Navn','Antall Lager',
    'Lagerplass','Sist Importert','Lev. ID'
    ]
    data['Utsolgt']['Daglig'].insert(0,colName)
    data['Utsolgt']['Daglig'].insert(0,[data['Tider']['yesterday']['weekday'].title() +' Uke-' + data['Tider']['yesterday']['weekNum']])
    data['Utsolgt']['Daglig'].insert(0,[data['Tider']['yesterday']['human']])
    data['Utsolgt']['Daglig'].insert(0,['Utsolgte Varer'])

    colName = [
    'Artikkel ID','Merke','Navn','Antall Solgt',
    'Dato','Klokketime','Pris','Rabatt','Betalingsmate'
    ]
    data['Salg']['Daglig'].insert(0,colName)
    data['Salg']['Daglig'].insert(0,[data['Tider']['yesterday']['weekday'].title() +' Uke-' + data['Tider']['yesterday']['weekNum']])
    data['Salg']['Daglig'].insert(0,[data['Tider']['yesterday']['human']])
    data['Salg']['Daglig'].insert(0,['Varesalg'])

    # loop through results in data and export spreadsheet
    for category in data:
        if category != 'Tider':
            for when in data[category]:
                if when == 'Daglig':
                    fileName = os.path.join(dirs[category],when,data[
                        'Tider']['yesterday']['YYYYMMDD']+'.xlsx')
                elif when == 'Ukentlig':
                    fileName = os.path.join(dirs[category],when,data[
                        'Tider']['yesterday']['YYYY-weekNum']+'.xlsx')
                elif when == 'Maanedlig':
                    fileName = os.path.join(dirs[category],when,data[
                        'Tider']['yesterday']['YYYYMMDD'][:6]+'.xlsx')
                else:
                    fileName == False

                if fileName != False:
                    exportXLSX(fileName)
                    Log(f'Spreadsheet: exporting {fileName}')






if __name__ == '__main__':
    data = {} # data for post and spreadsheet appended here
    post = {} # data only for post appended here

    initialize() # daily insert id_records in data warehouse if new data from store

    getRecords() # daily populate the data dict with new data from the store

    updateCIP() # daily send the newly populated data and insert into data warehouse

    writeSpreadsheet() # daily export spreadsheets from the data
