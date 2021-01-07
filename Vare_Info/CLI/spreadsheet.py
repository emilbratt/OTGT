
def exportXLSX(currentFile,data):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(list(row))

    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # set cloumn width
    if len(data[0]) < 27: # avoid exceeding the alphabet
        # apply default col width
        cellWidth = {}
        for i in range(len(data[0])):
            cellWidth[i] = 5
        # set width of col based on width of longest cell value
        for row in data:
            for i,cell in enumerate(row):
                if i not in cellWidth:
                    cellWidth[i] = 1
                try:
                    if cellWidth[i] < len(str(cell)):
                        cellWidth[i] = len(str(cell))
                except KeyError:
                    continue
        # apply width
        for i,key in enumerate(cellWidth):
            ws.column_dimensions[str(chr(65+i))].width = cellWidth[i]+5

    freeze = None  # find freeze point for titles while appending
    for i,row in enumerate(data):
        if len(row) > 2 and freeze == None:
            freeze = 'A'+str(i+2)
    if freeze != None:
        ws.freeze_panes = ws[freeze]


    wb.save(currentFile)
    return currentFile
